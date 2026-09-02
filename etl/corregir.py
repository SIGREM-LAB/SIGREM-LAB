"""Genera `corregido/` a partir de `original/`. El original NUNCA se toca.

    python -m etl.corregir              genera y verifica
    python -m etl.corregir --verificar  solo verifica lo ya generado

    # otra carpeta de entregas
    python -m etl.corregir --origen etl/Inventarios-JD2026 \
                           --destino etl/Inventarios-JD2026-corregido

Aqui van SOLO las correcciones deterministas: las que tienen un destino unico
y comprobable sin preguntarle a nadie. Una errata cuyo arreglo dependa de saber
cuantas piezas trae una caja, o de si ese frasco se peso o se midio, no entra
aqui y no debe entrar nunca: eso lo resuelve la pantalla de depuracion
(`docs/plans/2026-08-26-pantalla-depuracion-inventario.md`).

Por que un script y no editar el Excel a mano: las correcciones viven aqui, no
en el binario. Cuando N3 mande su siguiente version se vuelve a correr y el
diff de `correcciones-aplicadas.csv` se lee. Editar el .xlsx a mano deja el
archivo cambiado y ninguna forma de saber que se cambio.
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from etl.extract.formato import CAMPOS, HOJAS_DE_DATOS

RAIZ = Path(__file__).resolve().parents[1]
DATOS = RAIZ / "etl" / "Datos-Reales-JD2026"
ORIGINAL = DATOS / "original"
CORREGIDO = DATOS / "corregido"

BITACORA = "correcciones-aplicadas.csv"
COLUMNAS = ("archivo", "hoja", "celda", "campo", "antes", "despues", "motivo")


@dataclass(frozen=True)
class Correccion:
    """Un reemplazo literal sobre un campo, opcionalmente acotado a unas filas.

    `filas` no es una optimizacion: es el cinturon de seguridad. Un
    `«portatil» -> «portátil»` suelto sobre toda la hoja podria pegarle a un
    renglon que todavia no he mirado. Acotado a la fila que revise, no.
    """

    hoja: str
    campo: str
    buscar: str
    poner: str
    motivo: str
    filas: tuple[int, ...] = ()
    # Cuantas celdas se espera cambiar. La verificacion falla si no cuadra:
    # que el archivo cambie y esto siga en silencio es justo lo que no quiero.
    esperadas: int | None = None


@dataclass(frozen=True)
class Numerico:
    """Una celda que trae un numero escrito como texto. Regla 1."""

    hoja: str
    campo: str
    fila: int
    motivo: str


# ---------------------------------------------------------------------------
# Las correcciones de N3
# ---------------------------------------------------------------------------
# Cada bloque dice de donde salio. Sin eso, dentro de un mes nadie sabe si
# «geado -> grado» fue un hallazgo o un capricho.

# 1. Regla 5. El propio ETL nombra el destino: normalizar.mueble() rechaza
#    «Gabienete 301» con «¿quiso decir gabinete?». 74 renglones de Material.
MUEBLES = [
    Correccion("Material", "mueble", "Gabienete", "Gabinete",
               "Regla 5 · errata de mueble; el ETL propone el destino",
               esperadas=74),
]

# 2. Regla 3 / identidad del articulo. Estos cinco pares son IDENTICOS al
#    quitar acentos, mayusculas, espacios y puntuacion, asi que unificarlos no
#    puede fusionar dos articulos distintos: es el unico caso en que unificar
#    es seguro. Sin esto, `catalogo.resolver()` crea dos articulos para una
#    sola cosa, porque compara por igualdad exacta.
#
#    Los tres primeros se doblan a la forma mayoritaria. Los dos de Material
#    estan 1-1, asi que ahi manda la ortografia: «termómetro» y «portátil».
PARTIDOS = [
    Correccion("Reactivos", "sustancia", "sólido,presentación", "sólido, presentación",
               "Identidad del articulo · falta un espacio; 2 renglones lo escriben bien",
               filas=(222,), esperadas=1),
    Correccion("Reactivos", "sustancia", "sólido (anaerobios)", "sólido, (anaerobios)",
               "Identidad del articulo · se dobla a la forma mayoritaria (2 de 3)",
               filas=(371,), esperadas=1),
    Correccion("Reactivos", "sustancia", "McCoy´s", "McCoy's",
               "Identidad del articulo · acento agudo usado como apostrofo (3 de 4)",
               filas=(784,), esperadas=1),
    Correccion("Material", "articulo", "Termométro", "Termómetro",
               "Identidad del articulo · empate 1-1; manda la ortografia",
               filas=(233,), esperadas=1),
    Correccion("Material", "articulo", "Potenciómetro portatil", "Potenciómetro portátil",
               "Identidad del articulo · empate 1-1; manda la ortografia",
               filas=(232,), esperadas=1),
]

# 3. Erratas de palabra con destino unico. El criterio para que una entre aqui:
#    la palabra mal escrita aparece <=3 veces y la correcta cientos, asi que no
#    hay ambiguedad sobre que se quiso escribir.
#
#    NO estan aqui «piseta», «nitrito», «subnitrato», «anhídrido» ni
#    «tricloruro»: el barrido las marco por parecido, pero las cinco son
#    palabras reales y distintas de aquella a la que se parecen. Corregirlas
#    seria inventar.
ERRATAS = [
    Correccion("Reactivos", "sustancia", "geado", "grado",
               "Errata · «grado» aparece cientos de veces", filas=(323, 324, 325),
               esperadas=3),
    Correccion("Reactivos", "sustancia", "purea", "pureza",
               "Errata · «pureza» x423", filas=(390, 440), esperadas=2),
    Correccion("Reactivos", "sustancia", "puraza", "pureza",
               "Errata · «pureza» x423", filas=(485, 1040), esperadas=2),
    Correccion("Reactivos", "sustancia", "sólio", "sólido",
               "Errata · «sólido» x791", filas=(769,), esperadas=1),
    # Ergoesterol es ademas el caso de los 20 frascos: sin esta correccion se
    # queda como un articulo suelto al lado de los 23 «Ergosterol».
    Correccion("Reactivos", "sustancia", "Ergoesterol", "Ergosterol",
               "Errata · «Ergosterol» x23 en la misma hoja", filas=(610,),
               esperadas=1),
]

# 4. Los siete deletreos de «presentación». Van aparte porque son la misma
#    palabra siete veces, no siete hallazgos distintos.
PRESENTACION = [
    Correccion("Reactivos", "sustancia", mal, "presentación",
               "Errata · «presentación» x1046", filas=(fila,), esperadas=1)
    for mal, fila in (
        ("resentación", 226),
        ("pesentación", 403),
        ("oresentación", 474),
        ("presentacio", 492),
        ("prensentación", 756),
        ("presentacón", 972),
        ("presentasión", 986),
        ("presentacio", 1035),
    )
]

# 5. Regla 4. `validar()` ya dobla «SIN MARCA» a «Sin marca» sola, pero gana la
#    PRIMERA forma que ve en la corrida, que depende del orden de los archivos.
#    Dejarlo escrito aqui hace que el resultado no dependa de ese orden.
MARCAS = [
    Correccion(hoja, "marca", "SIN MARCA", "Sin marca",
               "Regla 4 · una sola grafia por marca")
    for hoja in ("Reactivos", "Insumos", "Material")
]

# 4b. «presentacion» sin acento, y «presentación» pegada al numero. Salieron al
#     revisar la Version 2 de N3 el 2 de septiembre de 2026; la Version 1 no
#     los traia o no se habian mirado.
#
#     Cumplen el mismo criterio que el bloque de arriba: destino unico, la
#     grafia correcta aparece mas de mil veces en la misma columna, y el
#     resultado no depende de saber nada que no este en la celda.
#
#     OJO con el orden de estas dos cadenas. «presentacion» contiene
#     «presentacio», asi que si esta correccion no estuviera acotada por filas
#     pisaria a la de arriba y dejaria «presentaciónn». Las filas son disjuntas
#     —492 y 1035 alli, estas diez aqui— y por eso no se tocan.
ACENTOS = [
    Correccion("Reactivos", "sustancia", "presentacion", "presentación",
               "Errata · «presentación» sin acento; x1046 bien escrita",
               filas=(142, 143, 144, 329, 366, 367, 711, 1028, 1029, 1030),
               esperadas=10),
]

# NO entran aqui las cinco «. Presentación 500 g» (filas 449, 450, 958, 960,
# 975). La mayuscula es correcta: van despues de un punto. Lo raro es el punto
# donde el resto de la hoja pone coma, y cambiar un separador es decidir por el
# almacen como se lee su articulo. Eso es de la pantalla de depuracion.
PEGADOS = [
    Correccion("Reactivos", "sustancia", "presentación250", "presentación 250",
               "Regla 3 · falta el espacio antes de la cantidad",
               filas=(489,), esperadas=1),
    Correccion("Reactivos", "sustancia", "presentación500", "presentación 500",
               "Regla 3 · falta el espacio antes de la cantidad",
               filas=(889,), esperadas=1),
]

CORRECCIONES = (MUEBLES + PARTIDOS + ERRATAS + PRESENTACION + ACENTOS
                + PEGADOS + MARCAS)

# 6. Regla 1. Un numero guardado como texto. `normalizar.numero()` lo rechaza y
#    con el se cae la hoja entera de Material.
#
#    Ojo: la fila 136 es tambien uno de los seis «Tapón» que estan en kg
#    mientras el resto esta en g. Eso NO se toca aqui: convertir 0.5 kg a 500 g
#    es exactamente la clase de decision que va a la pantalla de depuracion.
NUMERICOS = [
    Numerico("Material", "cantidad", 136,
             "Regla 1 · numero guardado como texto"),
]


@dataclass
class Bitacora:
    filas: list[dict[str, str]] = field(default_factory=list)

    def anotar(self, **campos: str) -> None:
        self.filas.append(campos)

    def a_csv(self, ruta: Path) -> None:
        # utf-8-sig por lo mismo que informe.py: este archivo lo abre gente en
        # Excel, y sin BOM los acentos salen rotos.
        with ruta.open("w", encoding="utf-8-sig", newline="") as f:
            escritor = csv.DictWriter(f, fieldnames=COLUMNAS)
            escritor.writeheader()
            escritor.writerows(self.filas)


def _letra(hoja: str, campo: str) -> str:
    """La columna sale de formato.CAMPOS, no de una constante propia.

    Si algun dia el formato mueve una columna, esto se mueve con el ETL en vez
    de corregir la celda equivocada en silencio.
    """
    try:
        return CAMPOS[hoja][campo]
    except KeyError as error:
        raise SystemExit(
            f"«{campo}» no es un campo de la hoja «{hoja}» en formato.CAMPOS") from error


def _aplicar(ws, correccion: Correccion, archivo: str,
             bitacora: Bitacora) -> int:
    letra = _letra(correccion.hoja, correccion.campo)
    filas = correccion.filas or range(1, ws.max_row + 1)
    cambiadas = 0
    for fila in filas:
        celda = ws[f"{letra}{fila}"]
        if not isinstance(celda.value, str) or correccion.buscar not in celda.value:
            continue
        antes = celda.value
        celda.value = antes.replace(correccion.buscar, correccion.poner)
        cambiadas += 1
        bitacora.anotar(archivo=archivo, hoja=correccion.hoja,
                        celda=f"{letra}{fila}", campo=correccion.campo,
                        antes=antes, despues=celda.value,
                        motivo=correccion.motivo)
    return cambiadas


def _aplicar_numerico(ws, numerico: Numerico, archivo: str,
                      bitacora: Bitacora) -> int:
    letra = _letra(numerico.hoja, numerico.campo)
    celda = ws[f"{letra}{numerico.fila}"]
    if not isinstance(celda.value, str):
        return 0
    antes = celda.value
    try:
        celda.value = float(antes.replace(",", "."))
    except ValueError:
        raise SystemExit(f"{archivo} {numerico.hoja}!{letra}{numerico.fila}: "
                         f"«{antes}» no se puede leer como numero")
    # «(número)» y no solo el valor: en el CSV, «0.5 -> 0.5» no deja ver que lo
    # que cambio fue el tipo, que es justo el motivo de la correccion.
    bitacora.anotar(archivo=archivo, hoja=numerico.hoja,
                    celda=f"{letra}{numerico.fila}", campo=numerico.campo,
                    antes=f"{antes} (texto)",
                    despues=f"{celda.value} (número)",
                    motivo=numerico.motivo)
    return 1


def corregir(origen: Path, destino: Path) -> Bitacora:
    """Copia el libro y le aplica las correcciones. Devuelve la bitacora."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    # copy2 y no una escritura nueva: si una hoja no lleva ninguna correccion,
    # el archivo de destino sigue siendo byte a byte el original.
    shutil.copy2(origen, destino)

    # data_only=False a proposito. Con True, openpyxl reemplaza cada formula
    # por su ultimo valor cacheado al guardar. Este libro no trae ninguna, pero
    # el siguiente almacen puede traerlas y el destrozo seria silencioso.
    libro = load_workbook(destino, data_only=False)
    bitacora = Bitacora()
    cuenta: dict[int, int] = {}

    for i, correccion in enumerate(CORRECCIONES):
        if correccion.hoja not in libro.sheetnames:
            continue
        cuenta[i] = _aplicar(libro[correccion.hoja], correccion,
                             destino.name, bitacora)

    for numerico in NUMERICOS:
        if numerico.hoja in libro.sheetnames:
            _aplicar_numerico(libro[numerico.hoja], numerico, destino.name,
                              bitacora)

    libro.save(destino)

    problemas = [
        f"  {c.hoja}·{c.campo} «{c.buscar}»: esperaba {c.esperadas}, cambio {cuenta[i]}"
        for i, c in enumerate(CORRECCIONES)
        if c.esperadas is not None and i in cuenta and cuenta[i] != c.esperadas
    ]
    if problemas:
        raise SystemExit("El archivo no es el que se reviso:\n" + "\n".join(problemas))

    return bitacora


# openpyxl serializa los numeros con «%.16g» y Excel guardaba hasta 17 cifras
# significativas, asi que 44.459999999999994 se reescribe 44.45999999999999. La
# deriva es de 4e-15 gramos: ruido de IEEE754 heredado de restar dos pesos en
# Excel, no un cambio de dato. Se tolera, pero se mide y se reporta: pasarlo por
# alto sin medirlo seria justo lo que este archivo existe para no hacer.
TOLERANCIA_RELATIVA = 1e-12


def _iguales(a, b) -> tuple[bool, float]:
    """Devuelve (si son el mismo valor, deriva relativa si son numeros)."""
    if isinstance(a, float) and isinstance(b, float):
        escala = max(abs(a), abs(b), 1.0)
        deriva = abs(a - b) / escala
        return deriva <= TOLERANCIA_RELATIVA, deriva
    return a == b, 0.0


def _vacia(v) -> bool:
    """Una celda sin dato.

    Excel distingue la cadena vacia de la celda en blanco; openpyxl, al
    reescribir el libro, colapsa la primera en la segunda. En N3 pasa una sola
    vez —Reactivos!I1066, una marca vacia— y es la misma clase de cosa que el
    ruido de coma flotante: artefacto de serializacion, no un dato que cambie.
    Se tolera, pero se cuenta y se reporta aparte; enterrarlo en la igualdad
    seria dejar de mirar justo donde este archivo existe para mirar.
    """
    return v is None or v == ""


def verificar(origen: Path, destino: Path,
              tocadas: set[tuple[str, str]]) -> tuple[list[str], list[str]]:
    """Que el corregido difiera del original SOLO en las celdas de la bitacora.

    Devuelve (fallas, notas). Una falla es motivo para tirar el archivo.

    Se comprueba celda por celda sobre TODO el libro, no solo sobre las columnas
    que el ETL mapea: una correccion que se lleve por delante una celda de la
    hoja «Reglas de captura» tiene que salir aqui. Y se comprueban las
    validaciones de datos —los desplegables del formato— porque si openpyxl las
    pierde, el archivo ya no sirve para devolverselo al almacen.
    """
    a = load_workbook(origen, data_only=False)
    b = load_workbook(destino, data_only=False)
    fallas: list[str] = []
    notas: list[str] = []

    if a.sheetnames != b.sheetnames:
        fallas.append(f"hojas: {a.sheetnames} -> {b.sheetnames}")
        return fallas, notas

    for nombre in a.sheetnames:
        ha, hb = a[nombre], b[nombre]

        if len(ha.merged_cells.ranges) != len(hb.merged_cells.ranges):
            fallas.append(f"{nombre}: celdas combinadas "
                          f"{len(ha.merged_cells.ranges)} -> "
                          f"{len(hb.merged_cells.ranges)}")
        va = len(ha.data_validations.dataValidation)
        vb = len(hb.data_validations.dataValidation)
        if va != vb:
            fallas.append(f"{nombre}: validaciones de datos {va} -> {vb}")

        # Una columna que se encoge solo es aceptable si estaba vacia. La
        # fantasma «N» de Material lo estaba; openpyxl no la reescribe.
        if hb.max_column < ha.max_column:
            perdidas = [
                c for c in range(hb.max_column + 1, ha.max_column + 1)
                if any(ha.cell(row=r, column=c).value is not None
                       for r in range(1, ha.max_row + 1))]
            if perdidas:
                fallas.append(f"{nombre}: se perdieron columnas con dato: {perdidas}")
            else:
                notas.append(f"{nombre}: se descarto/aron "
                             f"{ha.max_column - hb.max_column} columna(s) "
                             f"fantasma, sin una sola celda con dato")

        esperadas = ruido = vaciadas = 0
        deriva_max = 0.0
        for fila in range(1, ha.max_row + 1):
            for col in range(1, ha.max_column + 1):
                celda = ha.cell(row=fila, column=col)
                va_, vb_ = celda.value, hb.cell(row=fila, column=col).value
                if (nombre, celda.coordinate) in tocadas:
                    esperadas += 1
                    continue
                if _vacia(va_) and _vacia(vb_):
                    vaciadas += va_ != vb_
                    continue
                igual, deriva = _iguales(va_, vb_)
                if igual:
                    if deriva:
                        ruido += 1
                        deriva_max = max(deriva_max, deriva)
                    continue
                fallas.append(f"{nombre}!{celda.coordinate} cambio sin estar en "
                              f"la bitacora: {va_!r} -> {vb_!r}")

        notas.append(f"{nombre}: {esperadas} celdas corregidas, "
                     f"{ruido} con ruido de coma flotante "
                     f"(deriva maxima {deriva_max:.1e}), "
                     f"{vaciadas} cadena vacia → celda en blanco")

    return fallas, notas


def _tocadas(ruta: Path) -> set[tuple[str, str]]:
    """Las celdas que la bitacora dice haber cambiado."""
    if not ruta.exists():
        return set()
    with ruta.open(encoding="utf-8-sig", newline="") as f:
        return {(fila["hoja"], fila["celda"]) for fila in csv.DictReader(f)}


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Corrige lo determinista del origen.")
    p.add_argument("--verificar", action="store_true",
                   help="no regenera; solo comprueba lo que ya hay")
    p.add_argument("--origen", type=Path, default=ORIGINAL,
                   help=f"por omisión, {ORIGINAL.relative_to(RAIZ)}")
    p.add_argument("--destino", type=Path, default=CORREGIDO,
                   help=f"por omisión, {CORREGIDO.relative_to(RAIZ)}")
    args = p.parse_args(argv)

    origen_raiz, destino_raiz = args.origen.resolve(), args.destino.resolve()

    if not origen_raiz.is_dir():
        print(f"No existe {origen_raiz}", file=sys.stderr)
        return 2

    # El destino no puede colgar del origen: `rglob` volveria a encontrar los
    # corregidos de la corrida anterior y los trataria como originales, o sea
    # correcciones sobre correcciones sin que nadie se entere.
    if origen_raiz == destino_raiz or destino_raiz.is_relative_to(origen_raiz):
        print(f"El destino no puede estar dentro del origen: "
              f"{destino_raiz} ⊂ {origen_raiz}", file=sys.stderr)
        return 2

    total = fallas_totales = 0
    for origen in sorted(origen_raiz.rglob("*.xlsx")):
        if origen.name.startswith("~$"):
            continue
        destino = destino_raiz / origen.relative_to(origen_raiz)
        print(f"  {origen.relative_to(origen_raiz)}")

        if not args.verificar:
            bitacora = corregir(origen, destino)
            bitacora.a_csv(destino.parent / BITACORA)
            total += len(bitacora.filas)
            print(f"    {len(bitacora.filas)} celdas corregidas → {BITACORA}")

        if not destino.exists():
            print(f"    falta {destino.relative_to(destino_raiz)}", file=sys.stderr)
            return 1

        fallas, notas = verificar(origen, destino,
                                  _tocadas(destino.parent / BITACORA))
        for nota in notas:
            print(f"    ok · {nota}")
        for falla in fallas:
            print(f"    FALLA · {falla}", file=sys.stderr)
        fallas_totales += len(fallas)

    if not args.verificar:
        print(f"\n  {total} correcciones en total. El original no se toco.")
    if fallas_totales:
        print(f"  {fallas_totales} fallas: el corregido NO es fiable.",
              file=sys.stderr)
        return 1
    print("  El corregido difiere del original solo en las celdas de la bitacora.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
