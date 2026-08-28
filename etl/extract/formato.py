"""Lee una hoja del formato unificado.

El almacén sale del NOMBRE del archivo (`N3-Equipos.xlsx`), no de una columna:
decisión D1 del spec, Nivel 1 del Léeme del formato. La hoja se toma del libro
—hay exactamente una de datos— y se coteja contra el nombre del archivo, para
que un archivo mal rotulado no pase inadvertido.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Reactivos lleva tres filas de encabezado agrupado (los apartados de la
# NOM-005-STPS), así que sus títulos de columna van una fila más abajo.
FILA_ENCABEZADO = {"Reactivos": 9}
FILA_ENCABEZADO_DEFECTO = 8

HOJAS_DE_DATOS = ("Reactivos", "Insumos", "Material", "Equipos",
                  "Materia biológica", "Electrónica")

ALMACENES = ("N3", "N4", "LUM", "LE")

# Cuando el almacén no está en el nombre del archivo, está en el de la carpeta.
# Va como mapa explícito y no como heurística: «Almacén-Nivel-3» no se parece a
# «N3» por ninguna regla, y adivinarlo mal carga el inventario de un almacén en
# la ficha de otro sin lanzar un solo error.
CARPETAS = {
    "almacen-nivel-3": "N3",
    "almacen-nivel-4": "N4",
    "almacen-lum": "LUM",
    "almacen-le": "LE",
}

# Qué columna del formato es qué campo. Lo que no está aquí se ignora: «No.» es
# el consecutivo del Excel, y «7.1-a) Uso principal», «Zona de riesgo» y
# «Número de personas expuestas» son constantes por almacén que viven en la
# tabla `almacen` (decisión D4 del spec del 18 de agosto).
CAMPOS: dict[str, dict[str, str]] = {
    "Reactivos": {
        # Ojo con la C: el formato la titula «Anaquel», pero el esquema unificó
        # a `mueble`, y «Anaquel 2» es un VALOR de mueble igual que
        # «Gabinete 301» o «Refrigerador».
        "sub_ubicacion": "B", "mueble": "C", "repisa": "D", "fila_cajon": "E",
        "color": "F", "hoja_seguridad": "G", "sustancia": "H", "marca": "I",
        "presentacion": "J", "peso_vacio": "K", "peso_total": "L",
        "cantidad": "M", "unidad": "N",
        "solido": "P", "liquido": "Q", "gas": "R",
        "caracteristica_quimica": "S", "caracteristica_toxica": "T",
        "riesgo_salud": "U", "riesgo_reactividad": "V",
        "riesgo_inflamabilidad": "W", "peligro_especial": "X",
        "implica_peligro": "Y", "observaciones": "AB",
    },
    "Insumos": {
        "clasificacion": "B", "articulo": "C", "especificacion": "D",
        "marca": "E", "cantidad": "F", "unidad": "G", "presentacion": "H",
        "sub_ubicacion": "I", "mueble": "J", "repisa": "K", "fila_cajon": "L",
        "observaciones": "M",
    },
    "Equipos": {
        "articulo": "B", "marca": "C", "modelo": "D", "numero_serie": "E",
        "numero_inventario": "F", "sub_ubicacion": "G", "laboratorio": "H",
        "mueble": "I", "funcionamiento": "J", "fecha_chequeo": "K",
        "mantenimiento": "L", "observaciones": "M",
    },
    "Materia biológica": {
        "articulo": "B", "origen_especie": "C", "cantidad": "D", "unidad": "E",
        "presentacion": "F", "metodo_conservacion": "G", "temperatura": "H",
        "fecha_recoleccion": "I", "fecha_preparacion": "J",
        "responsable_muestra": "K", "sub_ubicacion": "L", "mueble": "M",
        "repisa": "N", "observaciones": "O",
    },
    "Electrónica": {
        "clasificacion": "B", "familia": "C", "articulo": "D",
        "especificacion": "E", "cantidad": "F", "unidad": "G",
        "presentacion": "H", "sub_ubicacion": "I", "mueble": "J",
        "coord_h": "K", "coord_v": "L", "coord_i": "M", "observaciones": "N",
    },
}
# Las dos hojas son idénticas entre sí.
CAMPOS["Material"] = CAMPOS["Insumos"]


class ArchivoIlegible(Exception):
    """El archivo no es una hoja del formato unificado, o no la que dice ser."""


@dataclass(frozen=True)
class Hoja:
    ruta: Path
    almacen: str
    nombre: str
    responsable: str | None
    periodo: str | None
    actualizado: str | None
    filas: tuple[int, ...]
    renglones: tuple[dict[str, Any], ...]

    @property
    def archivo(self) -> str:
        return self.ruta.name


def slug(nombre: str) -> str:
    """«Materia biológica» → «materia-biologica». Pública: la usa cargar.py."""
    d = unicodedata.normalize("NFD", nombre.lower())
    return "".join(c for c in d if unicodedata.category(c) != "Mn").replace(" ", "-")


def _extraer(ws, nombre: str) -> tuple[tuple[int, ...], tuple[dict[str, Any], ...]]:
    """Los renglones con dato de una hoja, y en qué fila de Excel está cada uno."""
    campos = CAMPOS[nombre]
    fila_enc = FILA_ENCABEZADO.get(nombre, FILA_ENCABEZADO_DEFECTO)

    filas: list[int] = []
    renglones: list[dict[str, Any]] = []
    for r in range(fila_enc + 1, ws.max_row + 1):
        crudo = {campo: ws[f"{letra}{r}"].value for campo, letra in campos.items()}
        if all(v is None for v in crudo.values()):
            continue
        filas.append(r)
        renglones.append(crudo)
    return tuple(filas), tuple(renglones)


def _hoja(ruta: Path, ws, almacen: str, nombre: str) -> Hoja:
    filas, renglones = _extraer(ws, nombre)
    return Hoja(
        ruta=ruta, almacen=almacen, nombre=nombre,
        responsable=ws["F4"].value, periodo=ws["B5"].value,
        actualizado=ws["F5"].value, filas=filas, renglones=renglones,
    )


def almacen_de(ruta: Path) -> str:
    """De qué almacén es el archivo.

    Sigue siendo la decisión D1 —el almacén sale del NOMBRE, no de una columna,
    porque ninguna hoja trae columna de almacén— pero el nombre puede ser el del
    archivo o el de su carpeta. Los archivos reales llegan como
    `Almacén-Nivel-3/Inventario final.xlsx`: el almacén está en la carpeta y el
    archivo se llama como al encargado le pareció.

    Se prueba, en orden: `N3.xlsx` (un libro por almacén), `N3-Equipos.xlsx`
    (un archivo por almacén-hoja, lo que usan los ejemplos) y la carpeta.
    """
    for candidato in (ruta.stem, ruta.parent.name):
        c = slug(candidato)
        for almacen in ALMACENES:
            if c == almacen.lower() or c.startswith(f"{almacen.lower()}-"):
                return almacen
        if c in CARPETAS:
            return CARPETAS[c]
    raise ArchivoIlegible(
        f"{ruta.name}: no se puede saber de qué almacén es. Nombra el archivo "
        f"«N3.xlsx» o ponlo en una carpeta reconocida ({', '.join(CARPETAS)})")


def leer(ruta: Path) -> Hoja:
    """Un archivo, una hoja de datos. El formato de los ejemplos."""
    almacen, _, resto = ruta.stem.partition("-")
    if not almacen or not resto:
        raise ArchivoIlegible(f"{ruta.name}: el nombre debe ser ALMACEN-Hoja.xlsx")

    libro = load_workbook(ruta, data_only=True)
    candidatas = [h for h in libro.sheetnames if h in HOJAS_DE_DATOS]
    if len(candidatas) != 1:
        raise ArchivoIlegible(
            f"{ruta.name}: se esperaba una hoja de datos y hay "
            f"{len(candidatas)}: {candidatas}")

    nombre = candidatas[0]
    if slug(nombre) != resto.lower():
        raise ArchivoIlegible(
            f"{ruta.name}: el archivo dice «{resto}» y la hoja es «{nombre}»")

    return _hoja(ruta, libro[nombre], almacen, nombre)


def leer_libro(ruta: Path, almacen: str | None = None) -> tuple[Hoja, ...]:
    """Un libro por almacén, con una hoja por clasificación.

    Es como entregan los almacenes de verdad: copian el formato unificado y
    llenan las hojas que les tocan. `leer()` no puede con eso —exige una sola
    hoja de datos y el nombre del archivo diciendo cuál es— y por eso el primer
    archivo real de N3 no se podía ni abrir.

    Las hojas salen en el orden de HOJAS_DE_DATOS y no en el del libro: el
    catálogo de artículos es global, así que el orden en que se cargan decide
    cuál renglón crea el artículo y cuál lo reutiliza. Que eso dependa de cómo
    ordenó sus pestañas el encargado haría la carga irreproducible.
    """
    libro = load_workbook(ruta, data_only=True)
    clave = almacen or almacen_de(ruta)

    presentes = [h for h in HOJAS_DE_DATOS if h in libro.sheetnames]
    if not presentes:
        raise ArchivoIlegible(
            f"{ruta.name}: no trae ninguna hoja de datos del formato "
            f"unificado. Se buscaron: {', '.join(HOJAS_DE_DATOS)}")

    return tuple(_hoja(ruta, libro[nombre], clave, nombre) for nombre in presentes)
